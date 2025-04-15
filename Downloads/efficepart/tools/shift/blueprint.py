#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
from datetime import datetime, timedelta
from io import BytesIO

# 現在のディレクトリと親ディレクトリをPythonパスに追加
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, '..', '..'))
sys.path.insert(0, parent_dir)

# お知らせ用のログ設定
import logging
logging.getLogger(__name__).info("Shift blueprint imported")

# サードパーティライブラリのインポート
from flask import Blueprint, request, render_template, send_file, current_app, jsonify
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# テンプレートとスタティックファイルのパスを設定
template_dir = os.path.join(current_dir, 'templates')
static_dir = os.path.join(current_dir, 'static')

# Blueprintの作成
shift_bp = Blueprint('shift', __name__, 
                    template_folder=template_dir,
                    static_folder=static_dir,
                    url_prefix='/shift')

# アップロード先のディレクトリを設定
UPLOAD_FOLDER = os.path.join(current_dir, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@shift_bp.route('/')
def index():
    """シフト作成ツールのメインページを表示"""
    return render_template('pages/shift_index.html')

@shift_bp.route('/generate', methods=['POST'])
def generate_shift():
    """シフト表を生成する"""
    try:
        # リクエストからデータを取得
        start_date = request.form.get('start_date')
        num_days = int(request.form.get('num_days', 7))
        num_workers = int(request.form.get('num_workers', 10))
        
        # Excelワークブックを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "シフト表"
        
        # スタイルの定義
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # ヘッダー行の作成
        ws.cell(row=1, column=1, value="名前").fill = header_fill
        
        # 日付ヘッダーの作成
        if start_date:
            current_date = datetime.strptime(start_date, '%Y-%m-%d')
        else:
            current_date = datetime.now()
            
        for i in range(num_days):
            date_str = (current_date + timedelta(days=i)).strftime('%m/%d (%a)')
            cell = ws.cell(row=1, column=i+2, value=date_str)
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(i+2)].width = 12
            
        # 名前行の作成
        for i in range(num_workers):
            cell = ws.cell(row=i+2, column=1, value=f"スタッフ{i+1}")
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            
        # シフトセルの作成
        for row in range(2, num_workers+2):
            for col in range(2, num_days+2):
                cell = ws.cell(row=row, column=col, value="")
                cell.border = border
                cell.alignment = center_alignment
                
        # エクセルファイルをバイナリストリームに保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # ファイルとして送信
        return send_file(
            output,
            as_attachment=True,
            download_name=f"shift_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except Exception as e:
        current_app.logger.error(f"シフト表生成中にエラーが発生しました: {str(e)}")
        return jsonify({"error": f"シフト表生成中にエラーが発生しました: {str(e)}"}), 500 