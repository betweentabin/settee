from flask import Flask, request, render_template, send_file
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from io import BytesIO
from openpyxl.utils import get_column_letter
import os

# テンプレートとスタティックファイルのパスを親ディレクトリに設定
template_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'templates', 'pages'))
static_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static'))
app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)

@app.template_filter('to_circle_number')
def to_circle_number(value):
    circle_numbers = [
        '①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩',
        '⑪', '⑫', '⑬', '⑭', '⑮', '⑯', '⑰', '⑱', '⑲', '⑳',
        '㉑', '㉒', '㉓', '㉔', '㉕', '㉖', '㉗', '㉘', '㉙', '㉚',
        '㉛', '㉜', '㉝', '㉞', '㉟', '㊱', '㊲', '㊳', '㊴', '㊵',
        '㊶', '㊷', '㊸', '㊹', '㊺', '㊻', '㊼', '㊽', '㊾', '㊿'
    ]
    return circle_numbers[value - 1] if 1 <= value <= 50 else str(value)

@app.route('/')
def index():
    return render_template('shift_index.html')

@app.route('/generate', methods=['POST'])
def generate():
    # ポジション関連パラメータ
    position_names = request.form.getlist('position_names')
    extra_positions = [pos for pos in request.form.getlist('extra_positions') if pos.strip()]  # 空の入力を除外
    
    # スタッフ基本情報
    staff_count = int(request.form['staff_count'])
    
    # スタッフ人数と必須ポジション数のチェック
    if len(position_names) > staff_count:
        return render_template(
            'shift_index.html',
            error="必須ポジション数がスタッフ人数を超えています。",
            form_data={
                'staff_count': staff_count,
                'position_names': position_names,
                'extra_positions': extra_positions,
                'work_start': request.form['work_start'],
                'work_end': request.form['work_end'],
                'break_start': request.form['break_start'],
                'break_end': request.form['break_end'],
                'break_duration': int(request.form['break_duration']),
                'exceptional_breaks': list(zip(
                    request.form.getlist('exceptional_break'),
                    request.form.getlist('exceptional_break_duration'),
                    request.form.getlist('exceptional_break_week')
                ))
            }
        )
    
    # 入力された臨時ポジションがある場合、合計ポジション数のチェック
    total_positions = len(position_names) + len(extra_positions)
    if total_positions > staff_count:
        return render_template(
            'shift_index.html',
            error="ポジション数の合計（必須 + 臨時）がスタッフ人数を超えています。",
            form_data={
                'staff_count': staff_count,
                'position_names': position_names,
                'extra_positions': extra_positions,
                'work_start': request.form['work_start'],
                'work_end': request.form['work_end'],
                'break_start': request.form['break_start'],
                'break_end': request.form['break_end'],
                'break_duration': int(request.form['break_duration']),
                'exceptional_breaks': list(zip(
                    request.form.getlist('exceptional_break'),
                    request.form.getlist('exceptional_break_duration'),
                    request.form.getlist('exceptional_break_week')
                ))
            }
        )

    work_start = datetime.strptime(request.form['work_start'], '%H:%M')
    work_end = datetime.strptime(request.form['work_end'], '%H:%M')
    
    # 休憩関連パラメータ
    break_start = datetime.strptime(request.form['break_start'], '%H:%M')
    break_end = datetime.strptime(request.form['break_end'], '%H:%M')
    base_break_duration = int(request.form['break_duration'])
    
    # 休憩時間帯が勤務時間内かチェック
    if break_start < work_start or break_end > work_end:
        return render_template(
            'shift_index.html',
            error="休憩時間帯は勤務時間内に設定してください。",
            # フォームの値を保持
            form_data={
                'staff_count': staff_count,
                'position_names': position_names,
                'extra_positions': extra_positions,
                'work_start': request.form['work_start'],
                'work_end': request.form['work_end'],
                'break_start': request.form['break_start'],
                'break_end': request.form['break_end'],
                'break_duration': base_break_duration,
                'exceptional_breaks': list(zip(
                    request.form.getlist('exceptional_break'),
                    request.form.getlist('exceptional_break_duration'),
                    request.form.getlist('exceptional_break_week')
                ))
            }
        )
    
    # 例外休憩取得
    exceptional_breaks = []
    exceptional_values = zip(
        request.form.getlist('exceptional_break'),
        request.form.getlist('exceptional_break_duration'),
        request.form.getlist('exceptional_break_week')
    )
    for eb, dur, week in exceptional_values:
        if eb == 'あり' and dur and week:
            exceptional_breaks.append({
                'duration': int(dur),
                'week': int(week)
            })

    shift_table = []
    current_time = work_start
    rotation_offset = 0
    staff_breaks = {i: None for i in range(staff_count)}
    staff_break_counts = {i: 0 for i in range(staff_count)}
    break_staff_count = max(0, staff_count - len(position_names))  # 休憩に回せる人数を固定

    while current_time < work_end:
        shift_row = {'time': current_time.strftime('%H:%M')}
        
        # 休憩終了チェック
        for i in range(staff_count):
            if staff_breaks[i] and current_time >= staff_breaks[i]:
                staff_breaks[i] = None

        # 個別休憩処理
        if break_start <= current_time < break_end:
            elapsed_minutes = (current_time - break_start).total_seconds() // 60
            # 現在休憩中のスタッフ数をカウント
            current_break_count = sum(1 for v in staff_breaks.values() if v is not None)
            
            # スタッフ総数から必要なポジション数を引いて、休憩可能人数を計算
            target_break_count = max(0, staff_count - len(position_names))
            
            if current_break_count < target_break_count:
                # 追加で休憩に入れる人数を計算
                available_break_slots = target_break_count - current_break_count
                
                if available_break_slots > 0:
                    start_idx = staff_count - rotation_offset - available_break_slots
                    end_idx = staff_count - rotation_offset
                    
                    targets = []
                    if start_idx < 0:
                        targets = list(range(max(staff_count + start_idx, 0), staff_count))
                        targets += list(range(0, max(end_idx, 0)))
                    else:
                        targets = range(max(start_idx, 0), min(end_idx, staff_count))
                    
                    for i in targets:
                        if 0 <= i < staff_count and staff_breaks[i] is None:
                            # スタッフの休憩回数に基づいて休憩時間を決定
                            current_break_duration = base_break_duration
                            current_week = staff_break_counts[i] + 1
                            for eb in exceptional_breaks:
                                if eb['week'] == current_week:
                                    current_break_duration = eb['duration']
                                    break
                            
                            break_end_time = current_time + timedelta(minutes=current_break_duration)
                            staff_breaks[i] = break_end_time
                            staff_break_counts[i] += 1
                    
                    rotation_offset = (rotation_offset + available_break_slots) % staff_count

        # スタッフ割り当て処理
        active_staff = [i for i in range(staff_count) if staff_breaks[i] is None]
        sorted_active_staff = sorted(active_staff, key=lambda x: (x - rotation_offset) % staff_count)
        
        position_assignments = {}
        # まず必要なポジションを確実に埋める
        for idx, staff_idx in enumerate(sorted_active_staff[:len(position_names)]):
            position_assignments[staff_idx] = position_names[idx]
        
        # 残りのスタッフを臨時ポジションに割り当て
        for idx, staff_idx in enumerate(sorted_active_staff[len(position_names):]):
            extra_idx = idx % len(extra_positions) if extra_positions else 0
            position_assignments[staff_idx] = extra_positions[extra_idx] if extra_positions else '待機'

        # シフト行作成
        for i in range(staff_count):
            status = '休憩' if staff_breaks[i] else position_assignments.get(i, '待機')
            shift_row[f'staff_{i+1}'] = status

        shift_table.append(shift_row)
        current_time += timedelta(minutes=15)

    return render_template(
        'shift_table.html',
        shift_table=shift_table,
        staff_count=staff_count,
        to_circle_number=to_circle_number
    )

@app.route('/download_excel', methods=['POST'])
def download_excel():
    # シフトデータとカラーマップの取得
    shift_data = request.json['shift_data']
    color_map = request.json['color_map']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "シフト表"
    
    # 中央揃えのスタイルを定義
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 薄いグレーの背景色を定義
    light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # 行の高さを設定
    ws.row_dimensions[1].height = 23  # 1行目（〇数字）
    ws.row_dimensions[2].height = 30  # 2行目（名前）を高く
    for row in range(3, len(shift_data) + 2):  # 3行目以降
        ws.row_dimensions[row].height = 23
    
    # 列幅の設定
    ws.column_dimensions['A'].width = 8  # 時間列
    for col in range(2, len(shift_data[0]) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15  # スタッフ列は少し広めに
    
    # 枠線のスタイルを定義（thin を dashed に変更）
    thin_border = Border(
        left=Side(style='dashed'),
        right=Side(style='dashed'),
        top=Side(style='dashed'),
        bottom=Side(style='dashed')
    )
    
    # 1行目用の枠線スタイル（左端用）
    first_row_left_border = Border(
        left=Side(style='medium'),
        right=Side(style='dashed'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    first_row_right_border = Border(
        left=Side(style='dashed'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    first_row_border = Border(
        left=Side(style='dashed'),
        right=Side(style='dashed'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    # 通常の行用の枠線スタイル
    thick_border = Border(
        left=Side(style='dashed'),
        right=Side(style='dashed'),
        top=Side(style='medium'),
        bottom=Side(style='dashed')
    )
    
    thick_right_border = Border(
        left=Side(style='dashed'),
        right=Side(style='medium'),
        top=Side(style='dashed'),
        bottom=Side(style='dashed')
    )
    
    # 最終行用の枠線スタイル
    last_row_left_border = Border(
        left=Side(style='medium'),
        right=Side(style='dashed'),
        top=Side(style='dashed'),
        bottom=Side(style='medium')
    )
    
    last_row_right_border = Border(
        left=Side(style='dashed'),
        right=Side(style='medium'),
        top=Side(style='dashed'),
        bottom=Side(style='medium')
    )
    
    last_row_border = Border(
        left=Side(style='dashed'),
        right=Side(style='dashed'),
        top=Side(style='dashed'),
        bottom=Side(style='medium')
    )
    
    # 太字のフォントスタイルを定義
    bold_font = Font(bold=True)
    
    staff_count = len(shift_data[0]) - 1
    total_rows = len(shift_data)
    
    # 1行目: 空白 + 〇数字
    first_row = [''] + [to_circle_number(i) for i in range(1, staff_count + 1)]
    for col, header in enumerate(first_row, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = center_alignment
        if col > 1:  # 最初の空白セル以外に背景色を設定
            cell.fill = light_gray_fill
        if col == 1:
            cell.border = first_row_left_border
        elif col == len(first_row):
            cell.border = first_row_right_border
        else:
            cell.border = first_row_border
    
    # 2行目以降: HTMLテーブルの現在のデータを使用
    for row_idx, row_data in enumerate(shift_data[1:], 2):
        time_value = row_data[0]
        is_hour_boundary = time_value.endswith(':00')
        is_last_row = row_idx == total_rows
        
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = center_alignment  # 中央揃えを適用
            
            # 2行目の場合、太字を適用
            if row_idx == 2:
                cell.font = bold_font
            
            # 枠線スタイルの決定
            if is_last_row:
                if col_idx == 1:
                    cell.border = last_row_left_border
                elif col_idx == len(row_data):
                    cell.border = last_row_right_border
                else:
                    cell.border = last_row_border
            else:
                if col_idx == 1:
                    cell.border = Border(
                        left=Side(style='medium'),
                        right=Side(style='medium'),
                        top=Side(style='medium') if is_hour_boundary else Side(style='dashed'),
                        bottom=Side(style='dashed')
                    )
                elif col_idx == len(row_data):
                    cell.border = Border(
                        left=Side(style='dashed'),
                        right=Side(style='medium'),
                        top=Side(style='medium') if is_hour_boundary else Side(style='dashed'),
                        bottom=Side(style='dashed')
                    )
                else:
                    cell.border = thick_border if is_hour_boundary else thin_border
            
            # シフトデータのセルの場合、色を設定
            if cell_value in color_map:
                color = color_map[cell_value].lstrip('#')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    # Excelファイルの保存
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='shift_table.xlsx'
    )

if __name__ == '__main__':
    app.run(debug=True)